import openpyxl
import exl_2_code_class

exl_file_name = "r1.xlsx"
wb = openpyxl.load_workbook(exl_file_name)
exl_file_sheet_count = len(wb.sheetnames)
exl_file_sheet_names = wb.sheetnames
object_list=[]

for i in range(0, exl_file_sheet_count):
    object_list.append(exl_2_code_class.exl_2_code_class(exl_file_name, exl_file_sheet_names[i]))

list_1=object_list[0].create_command()
print(list_1)




