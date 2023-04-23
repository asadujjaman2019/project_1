import openpyxl
class exl_2_code_class:
    def __init__(self,exl_file_name,exl_sheet_name):
        self.wb_object = openpyxl.load_workbook(exl_file_name)
        self.exl_sheet_name=exl_sheet_name
        self.ws_object=self.wb_object[self.exl_sheet_name]
        self.ws_row_count=self.ws_object.max_row
        self.ws_column_count=self.ws_object.max_column
        self.cell_values=[]
        self.commands_list=[]

    def create_command(self):

        for k in range(2, self.ws_row_count+1):
            for j in range(1,self.ws_column_count+1):
                self.cell_values.append(self.ws_object.cell(row=k, column=j).value)
            self.formate_command(self.cell_values)
            self.cell_values=list()
        return self.commands_list

    def formate_command(self,cell_value):
        self.interface = "interface" + cell_value[2]
        self.ip = "ip address " + cell_value[0] + " subnetmask " + cell_value[1]
        self.commands_list.append(self.interface)
        self.commands_list.append(self.ip)




