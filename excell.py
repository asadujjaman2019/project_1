import openpyxl
import excell_def

values=list()
file="r1.xlsx"
worksheet=['r1','r2']
wb = openpyxl.load_workbook(file)

for i in range(0,2):
    ws = wb[worksheet[i]]
    #print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    row=int(ws.max_row)
    column=int(ws.max_column)

    for k in range(2,ws.max_row+1):
        for j in range(1,ws.max_column + 1):
            values.append(ws.cell(row=k, column=j).value)
            #print("k=",k,values)

        excell_def.send_to_r(worksheet[i],values)
        values=list()

#print(ws.cell(row=3,column=3).value)
#values = [ws.cell(row=2,column=i).value for i in range(1,ws.max_column+1)]
#print(values[0])


#r1=['1.1.1.1','255.255.255.0','gig0/0']
#dict1= {"interface": "interface+ r1[2]", "ip": "ip address + r1[0] + subnetmask + r1[1]" }
#result = ssh.send_command(dict[interface])
#result = ssh.send_command(dict[ip])

